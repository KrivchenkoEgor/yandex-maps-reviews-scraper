"""
exporter — экспорт отзывов в Excel (3 листа), CSV и JSON.

Excel-структура из AGENTS.md:
- Лист 1: Отзывы (все поля + фильтры, форматирование)
- Лист 2: Статистика (распределение по рейтингам, тренды по месяцам, топ позитив/негатив)
- Лист 3: Магазин (инфо о магазине)
"""

import csv
import json
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from loguru import logger

from app import config

# ---------------------------------------------------------------------------
# Хелперы
# ---------------------------------------------------------------------------

def _ensure_output_dir(out_dir: str | Path | None = None) -> Path:
    p = Path(out_dir) if out_dir else Path(config.OUTPUT_DIR)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _safe_filename(name: str) -> str:
    import re
    s = re.sub(r'[\\/*?:"<>|]', "_", name).strip()
    return s[:80] or "export"


def to_excel_date(v: Any) -> Any:
    """ISO-строку 'YYYY-MM-DD' превратить в date — Excel покажет ДД.ММ.ГГГГ
    и будет сортировать/фильтровать как дату. Не-даты вернуть как есть."""
    if isinstance(v, (datetime, date)):
        return v
    s = str(v or "").strip()
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return v


def excel_date_sort_key(v: Any) -> str:
    """Ключ сортировки по дате: ISO-строка для настоящих дат, '' для остального
    (пустые и нестандартные значения уходят в конец при обратной сортировке)."""
    d = to_excel_date(v)
    return d.isoformat() if isinstance(d, date) else ""


def _reviews_sorted_newest_first(reviews: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Сортировка отзывов от самой новой даты к самой старой (не-даты — в конец)."""
    return sorted(reviews, key=lambda r: excel_date_sort_key(r.get("date")), reverse=True)


def _reviews_to_df(reviews: list[dict[str, Any]], excel: bool = False) -> pd.DataFrame:
    """Преобразовать список отзывов в DataFrame для экспорта.
    excel=True: даты — объектами date (отображение ДД.ММ.ГГГГ задаётся форматом ячейки)."""
    rows = []
    for r in reviews:
        owner = r.get("owner_response")
        owner_text = owner.get("text", "") if isinstance(owner, dict) else (owner or "")
        owner_date = owner.get("date", "") if isinstance(owner, dict) else ""
        rows.append({
            "ID": r.get("review_id", ""),
            "Автор": r.get("author", ""),
            "Рейтинг": r.get("rating", ""),
            "Дата": to_excel_date(r.get("date", "")) if excel else r.get("date", ""),
            "Текст": r.get("text", ""),
            "Фото (кол-во)": len(r.get("photos") or []),
            "Фото URL": "; ".join(r.get("photos") or []),
            "Ответ владельца": owner_text,
            "Дата ответа": to_excel_date(owner_date) if excel else owner_date,
            "Лайки": r.get("likes", 0),
            "Дизлайки": r.get("dislikes", 0),
            "Проверенный": "Да" if r.get("is_verified") else "Нет",
        })
    df = pd.DataFrame(rows)
    return df


def _stats_sheet(reviews: list[dict[str, Any]]) -> pd.DataFrame:
    """Статистика для листа 2."""
    total = len(reviews)
    ratings = [r.get("rating") for r in reviews if isinstance(r.get("rating"), int)]
    cnt = Counter(ratings)
    months = Counter()
    for r in reviews:
        d = r.get("date") or ""
        if re.match(r"^\d{4}-\d{2}", d):
            months[d[:7]] += 1

    # топ позитив/негатив по длине текста
    sorted_by_text = sorted(reviews, key=lambda x: len(x.get("text") or ""), reverse=True)
    top_text = (sorted_by_text[0].get("text") or "")[:120] + "..." if sorted_by_text else ""

    data = [
        ["Всего отзывов", total],
        ["Средний рейтинг", round(sum(ratings)/len(ratings), 2) if ratings else "—"],
        ["", ""],
        ["Распределение по рейтингам", ""],
        ["  5★", cnt.get(5, 0)],
        ["  4★", cnt.get(4, 0)],
        ["  3★", cnt.get(3, 0)],
        ["  2★", cnt.get(2, 0)],
        ["  1★", cnt.get(1, 0)],
        ["", ""],
        ["Тренд по месяцам (YYYY-MM)", "Кол-во"],
    ]
    for m in sorted(months):
        data.append([f"  {m}", months[m]])
    data += [
        ["", ""],
        ["Самый длинный отзыв (превью)", top_text],
    ]
    df = pd.DataFrame(data, columns=["Метрика", "Значение"])
    return df


def _shop_sheet(shop: dict[str, Any]) -> pd.DataFrame:
    data = [
        ["Название", shop.get("name", "")],
        ["Адрес", shop.get("address", "")],
        ["Рейтинг", shop.get("rating", "")],
        ["Всего отзывов (на Яндексе)", shop.get("total_reviews", "")],
        ["OID", shop.get("oid", "")],
        ["URL", shop.get("url", "")],
        ["Дата выгрузки", datetime.now().strftime("%d.%m.%Y %H:%M")],
    ]
    return pd.DataFrame(data, columns=["Поле", "Значение"])


# ---------------------------------------------------------------------------
# Публичные API
# ---------------------------------------------------------------------------

def export_excel(shop: dict[str, Any], reviews: list[dict[str, Any]], filename: str | None = None, out_dir: str | Path | None = None) -> Path:
    """
    Экспорт в Excel с 3 листами, форматированием и автофильтрами.
    Возвращает путь к файлу.
    """
    outp = _ensure_output_dir(out_dir)
    base = _safe_filename(shop.get("name") or shop.get("oid") or "shop")
    if filename is None:
        filename = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    path = outp / filename

    # Отзывы — от самой новой даты к самой старой
    reviews_sorted = _reviews_sorted_newest_first(reviews)
    df_reviews = _reviews_to_df(reviews_sorted, excel=True)
    df_stats = _stats_sheet(reviews)
    df_shop = _shop_sheet(shop)

    # Используем openpyxl для форматирования
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_reviews.to_excel(writer, sheet_name="Отзывы", index=False)
        df_stats.to_excel(writer, sheet_name="Статистика", index=False)
        df_shop.to_excel(writer, sheet_name="Магазин", index=False)

        # Форматирование через openpyxl
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=10)

            for sheet_name, df in [("Отзывы", df_reviews), ("Статистика", df_stats), ("Магазин", df_shop)]:
                ws = writer.sheets[sheet_name]
                # Заголовок
                for col in range(1, ws.max_column + 1):
                    c = ws.cell(row=1, column=col)
                    c.fill = header_fill
                    c.font = header_font
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                # Автофильтр только для Отзывов
                if sheet_name == "Отзывы" and df.shape[0] > 0:
                    ws.auto_filter.ref = ws.dimensions
                    ws.freeze_panes = "A2"
                    # Даты — формат ДД.ММ.ГГГГ (значения — настоящие даты Excel)
                    for idx, col_name in enumerate(df.columns, 1):
                        if col_name in ("Дата", "Дата ответа"):
                            for row in range(2, ws.max_row + 1):
                                ws.cell(row=row, column=idx).number_format = "DD.MM.YYYY"
                    # Ширины колонок
                    widths = {"ID": 14, "Автор": 18, "Рейтинг": 9, "Дата": 12, "Текст": 60, "Фото (кол-во)": 12, "Фото URL": 40, "Ответ владельца": 40, "Дата ответа": 12, "Лайки": 8, "Дизлайки": 9, "Проверенный": 12}
                    for idx, col_name in enumerate(df.columns, 1):
                        w = widths.get(col_name, 15)
                        ws.column_dimensions[get_column_letter(idx)].width = w
                    # Высота строк и перенос для текста
                    for row in ws.iter_rows(min_row=2):
                        ws.row_dimensions[row[0].row].height = 30
                        # Текст — перенос
                        # row[4] = Текст колонка (5-я)
                        if len(row) >= 5:
                            row[4].alignment = Alignment(wrap_text=True, vertical="center")
                else:
                    # Для статистики и магазина — автоширина
                    for idx, col in enumerate(df.columns, 1):
                        max_len = max([len(str(v)) for v in [col] + df[col].astype(str).tolist()] + [10])
                        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)
        except Exception as e:
            logger.warning(f"Форматирование Excel пропущено: {e}")

    logger.info(f"Excel сохранён: {path} ({len(reviews)} отзывов)")
    return path


def export_csv(shop: dict[str, Any], reviews: list[dict[str, Any]], filename: str | None = None, out_dir: str | Path | None = None) -> Path:
    outp = _ensure_output_dir(out_dir)
    base = _safe_filename(shop.get("name") or shop.get("oid") or "shop")
    if filename is None:
        filename = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    if not filename.endswith(".csv"):
        filename += ".csv"
    path = outp / filename
    df = _reviews_to_df(reviews)
    df.to_csv(path, index=False, encoding="utf-8-sig", quoting=csv.QUOTE_MINIMAL)
    logger.info(f"CSV сохранён: {path}")
    return path


def export_json(shop: dict[str, Any], reviews: list[dict[str, Any]], filename: str | None = None, out_dir: str | Path | None = None) -> Path:
    outp = _ensure_output_dir(out_dir)
    base = _safe_filename(shop.get("name") or shop.get("oid") or "shop")
    if filename is None:
        filename = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    if not filename.endswith(".json"):
        filename += ".json"
    path = outp / filename
    payload = {
        "shop": {
            "oid": shop.get("oid"),
            "name": shop.get("name"),
            "address": shop.get("address"),
            "rating": shop.get("rating"),
            "total_reviews": shop.get("total_reviews"),
            "url": shop.get("url"),
            "exported_at": datetime.now().isoformat(),
        },
        "reviews": reviews,
        "stats": {
            "count": len(reviews),
            "avg_rating": round(sum(r["rating"] for r in reviews if isinstance(r.get("rating"), int)) / max(1, len([r for r in reviews if isinstance(r.get("rating"), int)])), 2) if reviews else None,
        }
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info(f"JSON сохранён: {path}")
    return path


if __name__ == "__main__":
    import datetime as _dt

    demo_shop = {"oid": "1659941740", "name": "Тест Магнит", "address": "Красный проспект, 50", "rating": 4.3, "total_reviews": 247, "url": "https://yandex.ru/maps/-/CTwsUYyk"}
    demo_reviews = [
        # вход специально НЕ по порядку — export_excel должен отсортировать
        {"review_id": "a2", "author": "Мария С.", "rating": 2, "date": "2024-03-10", "text": "Очереди", "photos": [], "owner_response": None, "likes": 0, "dislikes": 1, "is_verified": False},
        {"review_id": "a3", "author": "Пётр", "rating": 4, "date": "", "text": "Норм", "photos": [], "owner_response": None, "likes": 3, "dislikes": 0, "is_verified": True},
        {"review_id": "a1", "author": "Иван П.", "rating": 5, "date": "2024-03-15", "text": "Отлично, свежие продукты", "photos": ["https://example.com/1.jpg"], "owner_response": {"text": "Спасибо!", "date": "2024-03-16"}, "likes": 12, "dislikes": 2, "is_verified": True},
    ]
    p1 = export_excel(demo_shop, demo_reviews, filename="demo_test.xlsx")
    p2 = export_csv(demo_shop, demo_reviews, filename="demo_test.csv")
    p3 = export_json(demo_shop, demo_reviews, filename="demo_test.json")
    print(f"✅ Excel: {p1} exists={p1.exists()}")
    print(f"✅ CSV: {p2} exists={p2.exists()}")
    print(f"✅ JSON: {p3} exists={p3.exists()}")

    # Сверка "было/стало" — количество строк
    import pandas as pd
    df = pd.read_excel(p1, sheet_name="Отзывы")
    assert len(df) == 3, f"ожидалось 3 строки, получили {len(df)}"

    # Порядок: от новой к старой (отзыв без даты — последним)
    from openpyxl import load_workbook
    wb = load_workbook(p1)
    ws = wb["Отзывы"]
    dates_col = 4  # "Дата"
    d1 = ws.cell(row=2, column=dates_col).value
    d2 = ws.cell(row=3, column=dates_col).value
    d3 = ws.cell(row=4, column=dates_col).value
    assert d1 == _dt.datetime(2024, 3, 15), f"первая строка должна быть 2024-03-15, получено {d1}"
    assert d2 == _dt.datetime(2024, 3, 10), f"вторая строка должна быть 2024-03-10, получено {d2}"
    assert d3 in ("", None), f"отзыв без даты должен быть последним, получено {d3!r}"
    print("✅ Сортировка: 15.03.2024 → 10.03.2024 → (без даты) OK")

    # Формат ячейки — русский ДД.ММ.ГГГГ
    fmt1 = ws.cell(row=2, column=dates_col).number_format
    fmt_answer = ws.cell(row=2, column=9).number_format  # "Дата ответа"
    assert fmt1 == "DD.MM.YYYY", f"формат даты: {fmt1}"
    assert fmt_answer == "DD.MM.YYYY", f"формат даты ответа: {fmt_answer}"
    print("✅ Формат дат в Excel: DD.MM.YYYY OK")

    p1.unlink(missing_ok=True); p2.unlink(missing_ok=True); p3.unlink(missing_ok=True)
    print("Все проверки exporter OK")
